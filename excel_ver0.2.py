# 製作：大平
# 注意：1文は80文字以内にすること。;で文を区切れる。

# import

import openpyxl as excel
import subprocess
import requests
from bs4 import BeautifulSoup
import datetime
import sys
from openpyxl.styles import numbers
import MySQLdb
import os
numbers.BUILTIN_FORMATS

# /import

# static変数

filePass = "python/repository/hello.xlsx"

# /static変数


# 処理内容

# webの処理

# URLのHTMLボディの取得
rec = requests.get(
    'https://stocks.finance.yahoo.co.jp/stocks/detail/?code=998407.O')

# HTMLボディのパース処理
class stock:
    def __init__(self,a_rate=None,title=None):
        self._rate=a_rate
        self._title=title



soup = BeautifulSoup(rec.text, 'html.parser')
stocks=[]

stock_details=soup.select("[class='ymuiEditLink mar0']")
rates=[]

for tmp1 in stock_details:
    tmp_soup=BeautifulSoup(f"""{tmp1}""", 'html.parser')
    tmp_soup.find('span').extract()
    rates.append(tmp_soup.text.strip().replace(',', ''))

stock_title=soup.select("dt.title")
titles=[]

for tmp1 in stock_title:
    tmp_soup=BeautifulSoup(f"""{tmp1}""", 'html.parser')
    tmp_soup.find('span').extract()
    titles.append(tmp_soup.text.strip())

 
print(rates)

print(titles)

r=0
for tmp1 in rates:
    stocks.append(stock(tmp1,titles[r]))
    r+=1

print(stocks)

today_index = soup.select_one(
    '#main > div.marB6.chartFinance.clearFix > div.innerDate > div:nth-child(1) > dl > dd > strong').text

print(today_index)

rate = today_index.replace(',', '')

print(rate)
# /webの処理

# フォルダとファイルの検索と新規作成
dir_path='python/repository'
excel_path='hello.xlsx'

is_there=os.path.isdir(dir_path)# os.path.exists(path)でもOK　boolean型で帰る

if is_there is False:
    print(f"{dir_path}フォルダがありません")
    os.makedirs(dir_path)
    wb=excel.Workbook()
    sheet=wb.active
    sheet.title='Sheet1'
    wb.save(filePass)
    

is_there=os.path.isfile(excel_path)# os.path.exists(path)でもOK　boolean型で帰る

if is_there is False:
    print(f"{excel_path}フォルダがありません")
    wb=excel.Workbook()
    sheet=wb.active
    sheet.title='Sheet1'
    wb.save(filePass)


# /フォルダとファイルの検索と新規作成


# sql

# 接続する
conn = MySQLdb.connect(user='root', passwd='Akira.299',
                       host='localhost', db='python')

cur = conn.cursor()

# SQL（データベースを操作するコマンド）を実行する
# userテーブルから、HostとUser列を取り出す
sql = "select Host,User from user"

cur.execute(sql)

# 実行結果を取得する
rows = cur.fetchall()

# 一行ずつ表示する
for row in rows:
#  print(row)
 pass

#データの追加


for stock in stocks:
    cur.execute(f'''insert into user(Host, User) values({stock._title},{stock._rate})''')

cur.close
conn.commit()
# 接続を閉じる
conn.close

# /sql


# excelの処理

# エクセルファイルのロード
excelBook = excel.load_workbook(filePass)

# シートのロード
sheet = excelBook['Sheet1']

# 現在日時の取得
dt_now = datetime.date.today()

if sheet.cell(row=2, column=2).value == None:
    sheet.cell(row=2, column=2).value = '日付'
    sheet.cell(row=2, column=3).value = '終値'

# 最終行の取得と新しい行番号処理

final_row =sheet.max_row
print(final_row)

date = sheet.cell(row=final_row, column=2).value

if date == str(dt_now):
    sys.exit()

print(sheet.cell(row=final_row, column=2).value)

newRowNumber = final_row+1

print(newRowNumber)

# セルの取得
cell = sheet.cell(row=newRowNumber, column=3)
date_cell = sheet.cell(row=newRowNumber, column=2)

# 値の書き込み
cell.value = float(rate)
date_cell.value = str(dt_now)

# セルの値のチェック
print(cell.value)

# セーブ
excelBook.save(filePass)

# /excelの処理

# /処理内容

# 起動して確認
EXCEL = r'C:/Users/akira\Documents/書類/python/hello.xlsx'
subprocess.Popen(['start', '', EXCEL], shell=True)
# /起動して確認
